from odoo import models, fields, api
import base64
import csv
import codecs
import base64
from io import BytesIO
import openpyxl
from odoo.exceptions import UserError

class HREmployee(models.Model):
    _inherit = "hr.employee"

    i_love_gb = fields.Boolean(string="I love GymBeam")
    salary = fields.Integer(string="Salary")
    tax = fields.Integer(string="Tax")
    total_salary = fields.Integer(string='Total Salary', compute='_compute_total_salary', store=True)

    @api.depends('salary', 'tax')
    def _compute_total_salary(self):
        for employee in self:
            employee.total_salary = employee.salary + employee.tax
    
    special_phone = fields.Char(string='Special Phone')
    employee_contacts = fields.Binary(string='Employee contacts')

    @api.model
    def create(self, vals):
        if 'special_phone' not in vals or not vals['special_phone']:
            vals['special_phone'] = "0901123456"
        return super(HREmployee, self).create(vals)

    def send_emails(self):
        
        if not self.employee_contacts:
            raise UserError("Please upload the Excel file in 'Employee Contacts' field.")

        
        # decoded_data = base64.b64decode(self.employee_contacts)
        # workbook = openpyxl.load_workbook(self.employee_contacts)
        # sheet = workbook.active
        email_list = []

        try:
            wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.employee_contacts)), read_only=True)
            ws = wb.active
            # decoded_data_str = decoded_data.decode('cp1252', errors='ignore')
            
            # lines = decoded_data_str.splitlines()
            # csv_data = csv.reader(lines, delimiter=',')
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                email_list.append({'email': row[0], 'subject': row[1]})

            # for row in csv_data:
            #     if len(row) < 2:
            #         continue
            #     if '\x00' in row:
            #         continue

            
                    

        except Exception as e:
            raise UserError(f"Error reading CSV data: {e}")

        if not email_list:
            raise UserError("No valid data found in the CSV file.")

        
        for email in email_list:
            mail = self.env['mail.mail'].create({
                'subject': email['subject'],
                'body_html': '<p>Welcome in GymBeam</p>',
                'email_to': email['email'],
            })
            mail.send()
    